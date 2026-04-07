import re
from collections import OrderedDict

import polars as pl

ns = "http://agricola.veronahe.no/"

# ─── Deck code → IRI mapping ─────────────────────────────────────────────────

DECK_CODE_TO_IRI = {
    "E":          ns + "deck_E",
    "I":          ns + "deck_I",
    "K":          ns + "deck_K",
    "Z":          ns + "deck_Z",
    "O":          ns + "deck_O",
    "Cz":         ns + "deck_Cz",
    "G":          ns + "deck_G",
    "G4":         ns + "deck_G4",
    "G5":         ns + "deck_G5",
    "G6":         ns + "deck_G6",
    "G7":         ns + "deck_G7",
    "G8":         ns + "deck_G8",
    "NL":         ns + "deck_NL",
    "Wm":         ns + "deck_Wm",
    "FL":         ns + "deck_FL",
    "WA":         ns + "deck_WA",
    "Pi":         ns + "deck_Pi",
    "BI":         ns + "deck_BI",
    "Fr":         ns + "deck_Fr",
    "Revised A":  ns + "deck_RevisedA",
    "Revised B":  ns + "deck_RevisedB",
    "Revised D":  ns + "deck_RevisedD",
    "Revised E":  ns + "deck_RevisedE",
    "MD1":        ns + "deck_MD1",
    "MD2":        ns + "deck_MD2",
    "MI":         ns + "deck_MI",
    "Unassigned": ns + "deck_Unassigned",
}


def deck_code_to_iri(code: str) -> str | None:
    """Convert a deck letter code to its full IRI."""
    if not code or not code.strip():
        return None
    return DECK_CODE_TO_IRI.get(code.strip())


# ─── Cost parsing helpers ─────────────────────────────────────────────────────

RESOURCE_ABBREV = OrderedDict([
    ("animal", "an"),
    ("boar",   "b"),
    ("cattle", "ca"),
    ("clay",   "c"),
    ("food",   "f"),
    ("grain",  "g"),
    ("reed",   "r"),
    ("sheep",  "sh"),
    ("stone",  "s"),
    ("vegetable", "v"),
    ("wood",   "w"),
])

COST_TOKEN_RE = re.compile(
    r'(\d+)\s*'
    r'(wild\s*boar|vegetable|cattle|sheep|grain|animal|stone|clay|reed|food|wood'
    r'|c|s|r|w|f|g|v|b)',
    re.IGNORECASE,
)

_ALIAS_TO_KEY = {
    "c": "clay", "s": "stone", "r": "reed", "w": "wood",
    "f": "food", "g": "grain", "v": "vegetable", "b": "boar",
    "wild boar": "boar", "wildboar": "boar",
    "clay": "clay", "stone": "stone", "reed": "reed", "wood": "wood",
    "food": "food", "grain": "grain", "vegetable": "vegetable",
    "cattle": "cattle", "sheep": "sheep", "animal": "animal",
    "boar": "boar",
}


def _parse_single_cost(cost_str: str) -> tuple[dict, str | None]:
    """Parse one cost alternative into ({resource: amount}, remainder)."""
    cost_str = cost_str.strip()
    if not cost_str:
        return {}, None

    parsed: dict[str, int] = {}
    spans: list[tuple[int, int]] = []

    for m in COST_TOKEN_RE.finditer(cost_str):
        amount = int(m.group(1))
        raw = m.group(2).lower().strip()
        key = _ALIAS_TO_KEY.get(raw, raw)
        parsed[key] = parsed.get(key, 0) + amount
        spans.append((m.start(), m.end()))

    remainder = cost_str
    for s, e in sorted(spans, reverse=True):
        remainder = remainder[:s] + remainder[e:]
    remainder = re.sub(r'[,\s]+', ' ', remainder).strip()
    return parsed, remainder or None


def _slug_for_alternative(parsed: dict, special: str | None) -> tuple[str, str]:
    tokens, labels = [], []
    for res in sorted(parsed, key=lambda r: RESOURCE_ABBREV.get(r, r)):
        abbr = RESOURCE_ABBREV.get(res, res[0])
        tokens.append(f"{parsed[res]}{abbr}")
        labels.append(f"{parsed[res]} {res}")
    if special:
        tokens.append(re.sub(r'[^a-zA-Z0-9]', '', special))
        labels.append(special)
    return "".join(tokens), ", ".join(labels)


def cost_to_iri(cost_str: str) -> dict | None:
    """Parse a cost string (may contain '/' alternatives) into IRI + metadata."""
    if not cost_str or not cost_str.strip():
        return None

    alternatives = re.split(r'\s*/\s*', cost_str.strip())
    slug_parts, label_parts = [], []
    primary_resources: dict | None = None
    special_iris: list[str] = []

    for alt in alternatives:
        parsed, special = _parse_single_cost(alt)
        if primary_resources is None:
            primary_resources = parsed.copy()
        slug, label = _slug_for_alternative(parsed, special)
        slug_parts.append(slug)
        label_parts.append(label)
        if special:
            special_iris.append(ns + re.sub(r'[^a-zA-Z0-9]', '', special))

    full_slug = "_or_".join(slug_parts)
    if not full_slug:
        return None

    return {
        "iri": ns + full_slug,
        "label": " / ".join(label_parts),
        "resources": primary_resources or {},
        "special_iri": special_iris[0] if special_iris else None,
    }


# ─── Card-text NLP extraction ────────────────────────────────────────────────

GAIN_PATTERNS = {
    "food":       re.compile(r'\b(?:receive|take|get|gives?\s+you|place)\b.*?\bfood\b|\bfood\b.*?\b(?:receive|take|get)\b|\bconvert\b.*?\bto\s+\d*\s*food\b|\b\d+\s*food\b', re.I),
    "wood":       re.compile(r'\b(?:receive|take|get)\b.*?\bwood\b|\bwood\b.*?\b(?:receive|take|get)\b|\b\d+\s*wood\b', re.I),
    "clay":       re.compile(r'\b(?:receive|take|get)\b.*?\bclay\b|\bclay\b.*?\b(?:receive|take|get)\b|\b\d+\s*clay\b', re.I),
    "stone":      re.compile(r'\b(?:receive|take|get)\b.*?\bstone\b|\bstone\b.*?\b(?:receive|take|get)\b|\b\d+\s*stone\b', re.I),
    "reed":       re.compile(r'\b(?:receive|take|get)\b.*?\breed\b|\breed\b.*?\b(?:receive|take|get)\b|\b\d+\s*reed\b', re.I),
    "grain":      re.compile(r'\b(?:receive|take|get|sow|plant)\b.*?\bgrain\b|\bgrain\b.*?\b(?:receive|take|get)\b|\b\d+\s*grain\b', re.I),
    "vegetable":  re.compile(r'\b(?:receive|take|get|sow|plant)\b.*?\bvegetable\b|\bvegetable\b.*?\b(?:receive|take|get)\b|\b\d+\s*vegetable\b', re.I),
    "sheep":      re.compile(r'\b(?:receive|take|get|buy)\b.*?\bsheep\b|\bsheep\b.*?\b(?:receive|take|get)\b|\b\d+\s*sheep\b', re.I),
    "boar":       re.compile(r'\b(?:receive|take|get|buy)\b.*?\b(?:wild\s*)?boar\b|\b(?:wild\s*)?boar\b.*?\b(?:receive|take|get)\b|\b\d+\s*(?:wild\s*)?boar\b', re.I),
    "cattle":     re.compile(r'\b(?:receive|take|get|buy)\b.*?\bcattle\b|\bcattle\b.*?\b(?:receive|take|get)\b|\b\d+\s*cattle\b', re.I),
    "room":       re.compile(r'\b(?:build|extend|add)\b.*?\broom\b|\broom\b.*?\b(?:build|add|extend)\b', re.I),
    "stable":     re.compile(r'\b(?:build|place)\b.*?\bstable\b|\bstable\b', re.I),
    "fence":      re.compile(r'\b(?:build|place)\b.*?\bfence\b|\bfence\b', re.I),
    "field":      re.compile(r'\b(?:plo(?:w|ugh)|field)\b', re.I),
    "renovation": re.compile(r'\brenovati?on\b', re.I),
    "bake":       re.compile(r'\bbake\s*(?:bread)?\b|\bbaking\b', re.I),
    "bonus_points": re.compile(r'\bbonus\s*point', re.I),
    "begging":    re.compile(r'\bbegging\s*card\b', re.I),
    "occupation": re.compile(r'\b(?:play|counts?\s+as)\b.*?\boccupation', re.I),
    "improvement": re.compile(r'\b(?:play|counts?\s+as)\b.*?\bimprovement', re.I),
    "family_growth": re.compile(r'\bfamily\s*growth\b|\boffspring\b|\bnewborn\b', re.I),
    "pasture":    re.compile(r'\bpasture\b', re.I),
    "action_space": re.compile(r'\baction\s*space\b', re.I),
    "cooking":    re.compile(r'\bcook(?:ing)?\b|\bconvert\b.*?\bto\s+food\b', re.I),
    "fireplace":  re.compile(r'\bfireplace\b', re.I),
    "cooking_hearth": re.compile(r'\bcooking\s*hearth\b', re.I),
}

AFFECT_PATTERNS = {
    "harvest":          re.compile(r'\bharvest\b', re.I),
    "feeding":          re.compile(r'\bfeed(?:ing)?\s*phase\b', re.I),
    "round_space":      re.compile(r'\bround\s*space\b', re.I),
    "sow":              re.compile(r'\bsow\b', re.I),
    "breeding":         re.compile(r'\bbreeding\b', re.I),
    "immediately":      re.compile(r'\bimmediately\b', re.I),
    "end_of_game":      re.compile(r'\bend\s*of\s*(?:the\s*)?game\b', re.I),
    "each_round":       re.compile(r'\beach\s*round\b|\bstart\s*of\s*(?:each|every)\s*round\b', re.I),
    "when_played":      re.compile(r'\bwhen\s*you\s*play\s*this\s*card\b', re.I),
    "whenever":         re.compile(r'\bwhenever\b', re.I),
    "once_per_round":   re.compile(r'\bonce\s*(?:per|each)\s*round\b|\bonce\s*during\b', re.I),
    "other_players":    re.compile(r'\bother\s*player\b|\ball\s*players\b|\bany\s*player\b', re.I),
    "minor_improvements": re.compile(r'\bminor\s*improvement', re.I),
    "major_improvements": re.compile(r'\bmajor\s*improvement', re.I),
}

# Known game terms that appear in quotes but are NOT card references
_NOT_CARD_REFS = {
    "sheep", "wild boar", "cattle", "grain", "vegetable", "wood", "clay",
    "stone", "reed", "food", "boar", "family growth", "fishing",
    "renovation", "bake bread", "travelling players", "traveling players",
    "sow and bake bread", "take 1 grain", "plough", "plow",
    "family growth without room", "newborn",
}

_KNOWN_CARD_RE = re.compile(
    r'\b(?:'
    r'Fireplace|Cooking\s*Hearth|Clay\s*Oven|Stone\s*Oven|Well|'
    r'Joinery|Pottery|Basketmaker\'?s?\s*Workshop|'
    r'Half-timbered\s*[Hh]ouse|Mansion|Corn\s*[Ss]torehouse|'
    r'Water\s*Mill|Windmill|Hand\s*Mill|'
    r'Chicken\s*[Cc]oop|Holiday\s*[Hh]ome|'
    r'Forest\s*Pasture|Bean\s*Field|Fruit\s*Tree|'
    r'Bread\s*Paddle|Bookshelf'
    r')\b',
    re.I,
)
_QUOTED_RE = re.compile(r"['\u201c\u201d\u2018\u2019\"]([A-Z][a-zA-Z\s\-]+?)['\u201c\u201d\u2018\u2019\"]")
_CARD_ID_REF_RE = re.compile(r'\b([EIKW][A-Z]?\d{2,5})\b')


def _normalise_ref(name: str) -> str:
    return re.sub(r'[^a-zA-Z0-9]', '', name.strip())


def extract_gains(text: str) -> list[str]:
    if not text:
        return []
    return sorted(k for k, pat in GAIN_PATTERNS.items() if pat.search(text))


def extract_affects(text: str) -> list[str]:
    if not text:
        return []
    return sorted(k for k, pat in AFFECT_PATTERNS.items() if pat.search(text))


def extract_relations(text: str) -> list[str]:
    """Extract references to other specific cards."""
    if not text:
        return []
    refs: set[str] = set()
    for m in _KNOWN_CARD_RE.finditer(text):
        refs.add(_normalise_ref(m.group(0)))
    for m in _QUOTED_RE.finditer(text):
        name = m.group(1).strip()
        if name.lower() not in _NOT_CARD_REFS and len(name) > 2:
            refs.add(_normalise_ref(name))
    for m in _CARD_ID_REF_RE.finditer(text):
        refs.add(m.group(1))
    return sorted(refs)


# ─── Helper: turn a list of strings into a comma-separated IRI string ────────

def _list_to_iris(items: list[str]) -> str:
    """Join items as comma-separated full IRIs."""
    if not items:
        return ""
    return ",".join(ns + item for item in items)


# ─── Main loader ─────────────────────────────────────────────────────────────

def _normalise_stats_name(name: str) -> str:
    """Normalise a card name for tournament-stats matching.

    Handles apostrophe variants (\u2019 vs '), removes spaces/punctuation,
    and lowercases so that truncated CSV names can match via prefix lookup.
    """
    return re.sub(r'[^a-z0-9]', '', name.lower())


def _load_tournament_stats(csv_path: str = "data/agricola_cards_all.csv") -> dict:
    """Load tournament stats from the legacy CSV, keyed by normalised card name.

    Returns two dicts: (exact_lookup, norm_lookup) where norm_lookup uses
    stripped alphanumeric keys for fuzzy/prefix matching.
    """
    import os
    if not os.path.exists(csv_path):
        return {}, {}
    stats_df = pl.read_csv(csv_path, infer_schema_length=10000)
    stats_df = stats_df.rename({c: c.replace(" ", "_") for c in stats_df.columns if " " in c})
    stats_cols = ["dealt", "drafted", "played", "won", "ADP",
                  "play_ratio", "win_ratio", "PWR", "PWR_no_log", "banned", "is_no"]
    exact_lookup = {}
    norm_lookup = {}
    for row in stats_df.iter_rows(named=True):
        name = (row.get("Name") or "").strip()
        vals = {col: row.get(col) for col in stats_cols}
        if name:
            exact_lookup[name.lower()] = vals
            norm_lookup[_normalise_stats_name(name)] = vals
    return exact_lookup, norm_lookup


def _load_database_xlsx(xlsx_path: str = "data/AgricolaCards_Database_260224_v2.xlsx") -> dict:
    """Load the curated card database XLSX (v2), keyed by card_uuid.

    Returns fields: PWRcorr, ADPcorr, Deck2, has_bonus_symbol,
    plus updated tournament stats (ADP, PWR, dealt, drafted, played, won).
    """
    import os
    if not os.path.exists(xlsx_path):
        # Fall back to v1 if v2 not present
        xlsx_path = "data/AgricolaCards_Database_260224.xlsx"
        if not os.path.exists(xlsx_path):
            return {}
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["cards"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0] if h is not None]
    lookup = {}
    for r in rows[1:]:
        d = {headers[i]: r[i] for i in range(len(headers))}
        uuid = d.get("card_uuid")
        if not uuid:
            continue
        # Normalise #N/A values
        def _clean(v):
            if v == "#N/A" or v == "N/A":
                return None
            return v
        def _float(v):
            v = _clean(v)
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        def _int(v):
            v = _clean(v)
            if v is None:
                return None
            try:
                return int(v)
            except (ValueError, TypeError):
                return None
        lookup[uuid] = {
            "PWRcorr": _float(d.get("PWRcorr")),
            "ADPcorr": _float(d.get("ADPcorr")),
            "ADP": _float(d.get("ADP")),
            "PWR": _float(d.get("PWR")),
            "Deck2": d.get("Deck2"),
            "has_bonus_symbol": bool(d.get("has_bonus_symbol")),
            "dealt": _int(d.get("dealt")),
            "drafted": _int(d.get("drafted")),
            "played": _int(d.get("played")),
            "won": _int(d.get("won")),
            "is_no": True,  # presence in the DB = Norwegian deck card
        }
    wb.close()
    return lookup


def _load_alt_images(agricola_json: str = "data/agricola.json") -> dict[str, str]:
    """Build a name→full_url map from agricola.json alt_image field."""
    import json as _json
    ALT_IMG_BASE = "https://hauk88.github.io/PlayAgricolaStatistics/img/"
    try:
        with open(agricola_json, "r") as f:
            data = _json.load(f)
        return {
            entry["name"].strip().lower(): ALT_IMG_BASE + entry["alt_image"]
            for entry in data
            if entry.get("alt_image")
        }
    except FileNotFoundError:
        return {}


# Local card images (name_key → URL path served by backend)
_LOCAL_CARD_IMAGES = {
    "resourcedealer": "/img/resource_dealer.jpg",
    "resourcecollector": "/img/resource_collector.jpg",
    "grainlover": "/img/grain_lover.jpg",
    "nilefarmer": "/img/nile_farmer.jpg",
    "firewoodcollector": "/img/firewood_collector.jpg",
    "porcelainmaker": "/img/porcelain_maker.jpg",
}


def load_cards(json_path: str = "data/cards.json",
               stats_csv: str = "data/agricola_cards_all.csv") -> pl.DataFrame:
    """Load cards from cards.json (single source of truth), enrich with
    tournament stats from the legacy CSV, merge curated XLSX data, and build IRIs."""
    import json as _json

    with open(json_path, "r") as f:
        raw_cards = _json.load(f)

    # Load tournament stats for merge (exact + normalised keys)
    stats_exact, stats_norm = _load_tournament_stats(stats_csv)

    # Load curated database XLSX (PWRcorr, Deck2, has_bonus_symbol)
    db_xlsx = _load_database_xlsx()

    # Load alt images from agricola.json (preferred over cards.json images)
    alt_images = _load_alt_images()

    # Build rows
    rows = []
    for c in raw_cards:
        name = c.get("card_name", "")
        name_lower = name.strip().lower()
        name_norm = _normalise_stats_name(name)
        # Try exact match first, then normalised (handles apostrophe variants),
        # then prefix match (handles truncated CSV names like "Animalhusbandryworke").
        st = stats_exact.get(name_lower)
        if not st:
            st = stats_norm.get(name_norm)
        if not st:
            # Prefix match: some CSV names are truncated at ~20 chars
            for csv_norm, csv_vals in stats_norm.items():
                if csv_norm and name_norm.startswith(csv_norm) and len(csv_norm) >= 10:
                    st = csv_vals
                    break
        if not st:
            st = {}
        xlsx = db_xlsx.get(c.get("card_uuid", ""), {})

        # Use first deck for primary deck column; keep all decks as comma-sep
        decks = c.get("decks") or []
        primary_deck = decks[0] if decks else ""
        primary_deck_iri = deck_code_to_iri(primary_deck)
        all_decks = ",".join(decks)

        # Cost string: new JSON uses "1W,1C" with commas → replace with spaces
        cost_raw = (c.get("cost") or "").replace(",", " ")

        # Image: local overrides → alt_image from agricola.json → cards.json
        name_key = name.strip().lower().replace(" ", "").replace("'", "").replace("-", "")
        local_url = _LOCAL_CARD_IMAGES.get(name_key)
        alt_url = alt_images.get(name_key)
        if not alt_url:
            alt_url = alt_images.get(name.strip().lower())
        imgs = c.get("card_image_urls") or []
        image_url = local_url or alt_url or (imgs[0] if imgs else None)

        # v2 xlsx values override CSV stats when present
        dealt = xlsx.get("dealt") if xlsx.get("dealt") is not None else st.get("dealt")
        drafted = xlsx.get("drafted") if xlsx.get("drafted") is not None else st.get("drafted")
        played = xlsx.get("played") if xlsx.get("played") is not None else st.get("played")
        won = xlsx.get("won") if xlsx.get("won") is not None else st.get("won")
        # Use 'is not None' checks to avoid dropping 0 values
        # Corrected values (PWRcorr/ADPcorr) are the primary stats;
        # fall back to raw PWR/ADP when corrected values are absent.
        _adp_corr = xlsx.get("ADPcorr")
        _adp_raw = xlsx.get("ADP") if xlsx.get("ADP") is not None else st.get("ADP")
        adp = _adp_corr if _adp_corr is not None else _adp_raw
        _pwr_corr = xlsx.get("PWRcorr")
        _pwr_raw = xlsx.get("PWR") if xlsx.get("PWR") is not None else st.get("PWR")
        pwr = _pwr_corr if _pwr_corr is not None else _pwr_raw

        # Recompute play_ratio and win_ratio from updated stats
        play_ratio = st.get("play_ratio")
        win_ratio = st.get("win_ratio")
        if dealt and played:
            try:
                play_ratio = float(played) / float(dealt)
            except (ValueError, ZeroDivisionError):
                pass
        if played and won:
            try:
                win_ratio = float(won) / float(played)
            except (ValueError, ZeroDivisionError):
                pass

        rows.append({
            "Card_ID": c.get("card_uuid", ""),
            "Name": name,
            "Type": c.get("card_type", ""),
            "Deck": primary_deck_iri,
            "DeckLabel": primary_deck,
            "Decks": all_decks,
            "Players": str(c.get("min_no_players") or "") if c.get("min_no_players") else None,
            "Cost": cost_raw or None,
            "Prerequisite": c.get("requirement") or None,
            "Card_Text": c.get("card_text") or None,
            "Bonus_Points": str(c.get("victory_points")) if c.get("victory_points") else None,
            "image_url": image_url,
            "card_creator": c.get("card_creator") or None,
            "is_passing_minor": c.get("is_passing_minor", False),
            # Tournament stats (v2 xlsx overrides CSV when present)
            "dealt": dealt,
            "drafted": drafted,
            "played": played,
            "won": won,
            "ADP": adp,
            "ADP_raw": _adp_raw,
            "play_ratio": play_ratio,
            "win_ratio": win_ratio,
            "PWR": pwr,
            "PWR_raw": _pwr_raw,
            "PWR_no_log": st.get("PWR_no_log"),
            "banned": st.get("banned"),
            # Norwegian deck: DB presence (by UUID) is the definitive source
            "is_no": xlsx.get("is_no", False),
            # Curated database XLSX fields (merged on card_uuid)
            "PWRcorr": pwr,   # now same as PWR (corrected is primary)
            "Deck2": xlsx.get("Deck2"),
            "has_bonus_symbol": xlsx.get("has_bonus_symbol", False),
        })

    df = pl.DataFrame(rows)

    # ── Subject IRI (use UUID) ────────────────────────────────────────────
    df = df.with_columns(
        (pl.lit(ns) + pl.col("Card_ID")).alias("subject")
    )

    # ── Type IRI (CamelCase) ──────────────────────────────────────────────
    df = df.with_columns(
        (pl.lit(ns) + pl.col("Type").str.replace(" ", "")).alias("Type")
    )

    # ── Cost IRI ──────────────────────────────────────────────────────────
    cost_col = df["Cost"].to_list()
    cost_iris, cost_labels = [], []
    cost_clay, cost_stone, cost_reed, cost_wood = [], [], [], []
    cost_food, cost_grain, cost_vegetable, cost_special = [], [], [], []

    for raw in cost_col:
        result = cost_to_iri(raw)
        if result is None:
            cost_iris.append(None); cost_labels.append(None)
            cost_clay.append(None); cost_stone.append(None)
            cost_reed.append(None); cost_wood.append(None)
            cost_food.append(None); cost_grain.append(None)
            cost_vegetable.append(None); cost_special.append(None)
        else:
            cost_iris.append(result["iri"])
            cost_labels.append(result["label"])
            res = result["resources"]
            cost_clay.append(res.get("clay")); cost_stone.append(res.get("stone"))
            cost_reed.append(res.get("reed")); cost_wood.append(res.get("wood"))
            cost_food.append(res.get("food")); cost_grain.append(res.get("grain"))
            cost_vegetable.append(res.get("vegetable"))
            cost_special.append(result["special_iri"])

    df = df.with_columns(
        pl.Series("hasCost", cost_iris, dtype=pl.Utf8),
        pl.Series("cost_label", cost_labels, dtype=pl.Utf8),
        pl.Series("cost_clay", cost_clay, dtype=pl.Int64),
        pl.Series("cost_stone", cost_stone, dtype=pl.Int64),
        pl.Series("cost_reed", cost_reed, dtype=pl.Int64),
        pl.Series("cost_wood", cost_wood, dtype=pl.Int64),
        pl.Series("cost_food", cost_food, dtype=pl.Int64),
        pl.Series("cost_grain", cost_grain, dtype=pl.Int64),
        pl.Series("cost_vegetable", cost_vegetable, dtype=pl.Int64),
        pl.Series("cost_special", cost_special, dtype=pl.Utf8),
    )

    return df


def build_card_annotations(cards_df: pl.DataFrame) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame]:
    """Build exploded DataFrames for gains, affects, and relations.

    Each row is a (subject, value) pair so maplib can produce one IRI triple per row.
    """
    texts = cards_df["Card_Text"].to_list()
    subjects = cards_df["subject"].to_list()
    gains_rows: list[dict] = []
    affects_rows: list[dict] = []
    relations_rows: list[dict] = []

    for subj, text in zip(subjects, texts):
        for g in extract_gains(text):
            gains_rows.append({"subject": subj, "gain": ns + g})
        for a in extract_affects(text):
            affects_rows.append({"subject": subj, "affect": ns + a})
        for r in extract_relations(text):
            relations_rows.append({"subject": subj, "relation": ns + r})

    gains_df = pl.DataFrame(gains_rows) if gains_rows else pl.DataFrame(schema={"subject": pl.Utf8, "gain": pl.Utf8})
    affects_df = pl.DataFrame(affects_rows) if affects_rows else pl.DataFrame(schema={"subject": pl.Utf8, "affect": pl.Utf8})
    relations_df = pl.DataFrame(relations_rows) if relations_rows else pl.DataFrame(schema={"subject": pl.Utf8, "relation": pl.Utf8})

    return gains_df, affects_df, relations_df


# ─── "Works Well With" combo inference ──────────────────────────────────────

def build_card_combos(
    cards_df: pl.DataFrame,
    gains_df: pl.DataFrame,
    affects_df: pl.DataFrame,
    relations_df: pl.DataFrame,
) -> pl.DataFrame:
    """Infer card synergy combos from structured card data.

    Produces (subject_a, subject_b, reason) rows.  Equivalent to the datalog
    rules we'd write for maplib once its parser is implemented:

      [?a, :worksWellWith, ?b] :- ...

    Combo rules implemented:
      1. Bidirectional card-text references
      2. Rare-resource supply chain (grain, vegetable cost)
      3. Animal husbandry (animal gain + breeding affect)
      4. Animal infrastructure (pasture/fence + animal gain)
      5. Grain-bake strategy (grain gain + bake gain with sow/harvest affect)
      6. Grain-sow field synergy (grain gain + field gain with sow affect)
      7. Family growth + room building
    """
    # ── Look-up sets ──────────────────────────────────────────────────────
    def _subjects_with_gain(g: str) -> set[str]:
        return set(gains_df.filter(pl.col("gain") == ns + g)["subject"].to_list())

    def _subjects_with_affect(a: str) -> set[str]:
        return set(affects_df.filter(pl.col("affect") == ns + a)["subject"].to_list())

    # Cards that cost a specific resource (> 0)
    def _subjects_costing(resource_col: str) -> set[str]:
        return set(
            cards_df.filter(pl.col(resource_col).is_not_null() & (pl.col(resource_col) > 0))
            ["subject"].to_list()
        )

    combos: set[tuple[str, str, str]] = set()

    def _add(a_set: set, b_set: set, reason: str):
        for a in a_set:
            for b in b_set:
                if a != b:
                    pair = tuple(sorted([a, b]))
                    combos.add((pair[0], pair[1], reason))

    # ── Rule 1: Bidirectional card-text references ────────────────────────
    # Resolve relation targets (name-based IRIs) to card subjects
    name_to_subject: dict[str, str] = {}
    for row in cards_df.iter_rows(named=True):
        norm = re.sub(r'[^a-zA-Z0-9]', '', row["Name"].strip())
        name_to_subject[norm] = row["subject"]

    for row in relations_df.iter_rows(named=True):
        src = row["subject"]
        target_local = row["relation"].replace(ns, "")
        target_subj = name_to_subject.get(target_local)
        if target_subj and src != target_subj:
            pair = tuple(sorted([src, target_subj]))
            combos.add((pair[0], pair[1], "card_reference"))

    # ── Rule 2: Rare-resource supply chain ────────────────────────────────
    # Grain producers → grain consumers (only 13 cards cost grain)
    _add(_subjects_with_gain("grain"), _subjects_costing("cost_grain"), "grain_supply")
    # Vegetable producers → vegetable consumers (only 4 cost vegetable)
    _add(_subjects_with_gain("vegetable"), _subjects_costing("cost_vegetable"), "vegetable_supply")

    # ── Rule 3: Animal + breeding (tight: animal card must ALSO gain pasture/fence/stable) ─
    breeders = _subjects_with_affect("breeding")
    animal_infra = _subjects_with_gain("pasture") | _subjects_with_gain("fence") | _subjects_with_gain("stable")
    for animal in ("sheep", "boar", "cattle"):
        # Tight combo: animal provider also builds infrastructure
        animal_infra_cards = _subjects_with_gain(animal) & animal_infra
        _add(animal_infra_cards, breeders, "animal_breeding")
        # Also: breeding cards that themselves gain animals
        breeder_animals = breeders & _subjects_with_gain(animal)
        animal_only = _subjects_with_gain(animal) - breeders
        _add(animal_only, breeder_animals, "animal_breeding")

    # ── Rule 4: Grain + bake (bake card must also affect sow or harvest) ──
    bakers = _subjects_with_gain("bake")
    sow_or_harvest = _subjects_with_affect("sow") | _subjects_with_affect("harvest")
    strategic_bakers = bakers & sow_or_harvest
    _add(_subjects_with_gain("grain"), strategic_bakers, "baking_strategy")

    # ── Rule 5: Family growth + room (tight: room card must also affect immediately) ─
    room_cards = _subjects_with_gain("room") & _subjects_with_affect("immediately")
    _add(_subjects_with_gain("family_growth"), room_cards, "family_room")

    # ── Rule 6: Multi-signal overlap ──────────────────────────────────────
    # Cards that share 2+ uncommon gain categories → strong synergy
    uncommon_gains = {"renovation", "cooking_hearth", "fireplace", "cooking",
                      "pasture", "fence", "stable", "begging"}
    subj_uncommon: dict[str, set[str]] = {}
    for row in gains_df.iter_rows(named=True):
        g = row["gain"].replace(ns, "")
        if g in uncommon_gains:
            subj_uncommon.setdefault(row["subject"], set()).add(g)
    # Find pairs sharing 2+ uncommon gains
    subjects_with_multi = [s for s, gs in subj_uncommon.items() if len(gs) >= 2]
    for i, a in enumerate(subjects_with_multi):
        for b in subjects_with_multi[i+1:]:
            shared = subj_uncommon[a] & subj_uncommon[b]
            if len(shared) >= 2:
                pair = tuple(sorted([a, b]))
                combos.add((pair[0], pair[1], "multi_signal"))

    # Build DataFrame
    rows = [{"subject_a": a, "subject_b": b, "combo_reason": r} for a, b, r in combos]
    if not rows:
        return pl.DataFrame(schema={"subject_a": pl.Utf8, "subject_b": pl.Utf8, "combo_reason": pl.Utf8})
    return pl.DataFrame(rows)


# ─── Cost permutation DataFrame (for the CostPermutation RDF instances) ──────

def build_cost_permutations(cards_df: pl.DataFrame) -> pl.DataFrame:
    """Build a DataFrame of unique CostPermutation resources for the template."""
    rows = []
    seen = set()

    for row in cards_df.iter_rows(named=True):
        iri = row.get("hasCost")
        if not iri or iri in seen:
            continue
        seen.add(iri)
        rows.append({
            "subject": iri,
            "label": row.get("cost_label", ""),
            "clay": row.get("cost_clay"),
            "stone": row.get("cost_stone"),
            "reed": row.get("cost_reed"),
            "wood": row.get("cost_wood"),
            "food": row.get("cost_food"),
            "grain": row.get("cost_grain"),
            "vegetable": row.get("cost_vegetable"),
            "costValue": row.get("cost_special"),
        })

    return pl.DataFrame(rows)


cards = load_cards()
cost_permutations = build_cost_permutations(cards)
card_gains, card_affects, card_relations = build_card_annotations(cards)
card_combos = build_card_combos(cards, card_gains, card_affects, card_relations)

# Columns the OTTR Card template does NOT know about – drop before maplib mapping
_EXTRA_COLS = {"image_url", "card_creator", "is_passing_minor", "Decks",
               "PWRcorr", "Deck2", "has_bonus_symbol", "PWR_raw", "ADP_raw",
               "DeckLabel"}
cards_for_rdf = cards.drop([c for c in _EXTRA_COLS if c in cards.columns])
